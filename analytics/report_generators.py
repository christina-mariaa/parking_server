import openpyxl
from io import BytesIO


def generate_xlsx_report(data):
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    if default_sheet.title == "Sheet":
        wb.remove(default_sheet)

    if 'statistics' in data:  # Статистика
        ws = wb.create_sheet(title="Статистика")
        ws.append(["Метрика", "Значение"])
        for key, value in data['statistics'].items():
            ws.append([key, value])
    if 'bookings' in data:  # Бронирования
        ws = wb.create_sheet(title="Бронирования")
        ws.append(["ID", "Email пользователя", "Дата начала", "Дата окончания", "Номер автомобиля", "Статус", "Тариф", "Парковочное место"])
        for booking in data['bookings']:
            ws.append([
                booking['id'], booking['car__user__email'], booking['start_time'], booking['end_time'],
                booking['car__license_plate'],
                booking['status'], booking['tariff__name'], booking['parking_place__spot_number']
            ])
    if 'payments' in data:  # Оплаты
        ws = wb.create_sheet(title="Оплаты")
        ws.append(["ID", "Сумма", "Дата оплаты", "ID бронирования", "Email пользователя"])
        for payment in data['payments']:
            ws.append([
                payment['id'], payment['amount'], payment['payment_date'],
                payment['booking_id'], payment['booking__car__user__email']
            ])
    if 'new_users' in data:  # Новые пользователи
        ws = wb.create_sheet(title="Новые пользователи")
        ws.append(["ID", "Email", "Имя", "Фамилия", "Дата регистрации"])
        for user in data['new_users']:
            ws.append([
                user['id'], user['email'], user['first_name'], user['last_name'], user['date_joined']
            ])
    if 'new_cars' in data:  # Новые автомобили
        ws = wb.create_sheet(title="Новые автомобили")
        ws.append(["ID", "Email пользователя", "Номер автомобиля", "Марка", "Модель", "Цвет", "Дата регистрации", "Удален"])
        for car in data['new_cars']:
            ws.append([
                car['id'], car['user__email'], car['license_plate'], car['make'],
                car['model'], car['color'],  car['registered_at'], car['is_deleted']
            ])
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer
